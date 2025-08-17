from openpyxl import load_workbook

from config import MAPEAMENTO_COLUNAS
from utils import copiar_linha_com_formula, filtrar_linhas, obter_ultima_linha_com_dados


def abrir_planilhas():
    """Abre planilha de origem e destino, retornando workbooks e worksheets."""
    wb_origem = load_workbook("Planilha_Filtro_Extracao.xlsx")
    ws_origem = wb_origem.active

    wb_destino = load_workbook("Planilha_Filtro_Acompanhamento.xlsx")
    ws_destino_geral = wb_destino["Acompanhamento"]
    ws_destino_ri = wb_destino["RI"]
    ws_destino_projetos = wb_destino["Projetos"]

    return (
        wb_origem,
        ws_origem,
        wb_destino,
        ws_destino_geral,
        ws_destino_ri,
        ws_destino_projetos,
    )


def preparar_mapeamento(ws_origem, ws_destino, mapa_colunas=None):
    """Prepara índices e colunas para cópia entre origem e destino."""
    cabecalho_origem = [cell.value for cell in ws_origem[1]]
    cabecalho_destino = [cell.value for cell in ws_destino[1]]

    indice_origem = {name: idx + 1 for idx, name in enumerate(cabecalho_origem)}
    indice_destino = {name: idx + 1 for idx, name in enumerate(cabecalho_destino)}

    # Se não passar mapa_colunas, assume colunas iguais
    if mapa_colunas is None:
        mapa_colunas = {
            col: col for col in cabecalho_destino if col in cabecalho_origem
        }

    # Colunas do destino que devem ser copiadas
    colunas_para_copiar = [
        indice_destino[dest] for dest in mapa_colunas.keys() if dest in indice_destino
    ]

    return mapa_colunas, indice_origem, indice_destino, colunas_para_copiar


def preparar_destino(ws_destino, linha_modelo: int = 2):
    """Limpa a planilha de destino, mantendo a linha modelo."""
    if ws_destino.max_row > linha_modelo:
        ws_destino.delete_rows(linha_modelo + 1, ws_destino.max_row - linha_modelo)
    print(f"Após limpeza, '{ws_destino.title}' tem {ws_destino.max_row} linhas")


def copiar_para_aba(
    ws_destino,
    linha_modelo,
    linhas_origem,
    ws_origem,
    mapa_colunas,
    indice_origem,
    indice_destino,
    colunas_para_copiar,
    ajustar_formulas: bool = True,
    colunas_extras: list[str] = None,
):
    """Copia linhas da origem para a aba de destino mantendo formatação e fórmulas."""
    linha_destino = 2

    for linha_origem in linhas_origem:
        copiar_linha_com_formula(
            ws_destino,
            linha_origem=linha_modelo,
            linha_destino=linha_destino,
            colunas=colunas_para_copiar,
            colunas_extras=colunas_extras if colunas_extras else [],
            ajustar_formulas=ajustar_formulas,
        )

        for coluna_destino, coluna_origem in mapa_colunas.items():
            idx_dest = indice_destino.get(coluna_destino)
            idx_origem = indice_origem.get(coluna_origem)
            if idx_dest is None or idx_origem is None:
                continue

            valor_origem = ws_origem.cell(row=linha_origem, column=idx_origem).value
            ws_destino.cell(row=linha_destino, column=idx_dest).value = valor_origem

        linha_destino += 1

    return linha_destino - 2  # total de registros copiados


def obter_ultima_linha(ws, coluna_chave):
    """
    Retorna o índice da última linha com dados em uma coluna chave.

    Args:
        ws: Worksheet (aba do Excel)
        coluna_chave (str): Nome do cabeçalho da coluna usada como referência.

    Returns:
        int: Número da última linha com dados.
    """
    cabecalho = [cell.value for cell in ws[1]]
    if coluna_chave not in cabecalho:
        raise ValueError(
            f"Coluna '{coluna_chave}' não encontrada no cabeçalho da aba {ws.title}"
        )

    idx_col = cabecalho.index(coluna_chave) + 1
    return obter_ultima_linha_com_dados(ws, idx_col)


def processar_acompanhamento(ws_origem, ws_destino, ultima_linha=None):
    # Limpar aba de destino
    preparar_destino(ws_destino)

    mapa_colunas, indice_origem, indice_destino, colunas_para_copiar = (
        preparar_mapeamento(ws_origem, ws_destino, MAPEAMENTO_COLUNAS)
    )

    linhas = list(range(2, ultima_linha + 1))
    total_copiados = copiar_para_aba(
        ws_destino,
        linha_modelo=2,
        linhas_origem=linhas,
        ws_origem=ws_origem,
        mapa_colunas=mapa_colunas,
        indice_origem=indice_origem,
        indice_destino=indice_destino,
        colunas_para_copiar=colunas_para_copiar,
        colunas_extras=["G", "H"],
    )
    print(f"Total de registros copiados para Acompanhamento: {total_copiados}")


def processar_ri(ws_origem, ws_destino, ultima_linha=None):
    # Limpar aba de destino
    preparar_destino(ws_destino)

    mapa_colunas, indice_origem, indice_destino, colunas_para_copiar = (
        preparar_mapeamento(ws_origem, ws_destino, MAPEAMENTO_COLUNAS)
    )

    # Copiar para RI (filtrando status != 'Fechado')
    linhas_para_RI = filtrar_linhas(
        ws_origem,
        indice_origem["Status"],
        excluir=["Fechado"],
        linha_inicial=2,
        linha_final=ultima_linha,
    )
    total_copiados = copiar_para_aba(
        ws_destino,
        linha_modelo=2,
        linhas_origem=linhas_para_RI,
        ws_origem=ws_origem,
        mapa_colunas=mapa_colunas,
        indice_origem=indice_origem,
        indice_destino=indice_destino,
        colunas_para_copiar=colunas_para_copiar,
        colunas_extras=["G", "H"],
    )
    print(f"Total de registros copiados para RI: {total_copiados}")


def processar_projetos(ws_origem, ws_destino, ultima_linha=None):
    # Limpar aba de destino
    preparar_destino(ws_destino)

    # prepara mapeamento entre origem e destino
    mapa_colunas, indice_origem, indice_destino, colunas_para_copiar = (
        preparar_mapeamento(ws_origem, ws_destino)
    )

    linhas = list(range(2, ultima_linha + 1))
    total_copiados = copiar_para_aba(
        ws_destino,
        linha_modelo=2,
        linhas_origem=linhas,
        ws_origem=ws_origem,
        mapa_colunas=mapa_colunas,
        indice_origem=indice_origem,
        indice_destino=indice_destino,
        colunas_para_copiar=colunas_para_copiar,
        ajustar_formulas=False,
    )
    print(f"Total de registros copiados para Projetos: {total_copiados}")


def main():
    # Abrir planilhas
    (
        wb_origem,
        ws_origem,
        wb_destino,
        ws_destino_geral,
        ws_destino_ri,
        ws_destino_projetos,
    ) = abrir_planilhas()
    print(f"Planilha de destino: {wb_destino.sheetnames}")

    # Aqui escolhemos qual coluna serve de referência
    ultima_linha = obter_ultima_linha(ws_origem, "TicketID")

    # Executa os processos.
    processar_acompanhamento(ws_origem, ws_destino_geral, ultima_linha)
    processar_ri(ws_origem, ws_destino_ri, ultima_linha)
    processar_projetos(ws_origem, ws_destino_projetos, ultima_linha)

    # Salvar planilha
    wb_destino.save("Planilha_Filtro_Acompanhamento_v2.xlsx")
    print("Planilha salva com sucesso.")

    wb_origem.close()
    wb_destino.close()


if __name__ == "__main__":
    main()

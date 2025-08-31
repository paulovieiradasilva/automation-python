import re
import time
from pathlib import Path
from copy import copy
from contextlib import contextmanager
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet


def log(message):
    """Função para registrar mensagens de log."""
    print(f"[LOG] {message}")


@contextmanager
def log_tempo(mensagem="Processo finalizado"):
    """Context manager para registrar o tempo de execução o de um bloco de
    código. Registra a mensagem de início e fim do bloco com o tempo de
    execução em minutos e segundos.

    Args:
        mensagem (str): Mensagem para identificar o bloco de código.
    """
    inicio = time.time()
    log(f"{mensagem} iniciado...")
    try:
        yield
    finally:
        fim = time.time()
        total = fim - inicio
        minutos, segundos = divmod(total, 60)
        log(f"{mensagem} em {int(minutos)} minutos e {segundos:.2f} segundos.")


def preparar_destino(ws_destino, linha_modelo: int = 2):
    """Limpa a planilha de destino, mantendo a linha modelo."""

    if ws_destino.max_row > linha_modelo:
        ws_destino.delete_rows(linha_modelo + 1, ws_destino.max_row - linha_modelo)
    log(
        f"[RELATÓRIO] Após limpeza, '{ws_destino.title}' tem {ws_destino.max_row} linhas"
    )


def preparar_pasta(subpasta: str | None = None) -> Path:
    """
    Garante que a estrutura de pastas de extração exista:
    - Sempre cria a pasta base 'data' no diretório do script.
    - Se 'subpasta' for informado, cria dentro de 'data'.

    Retorna o Path absoluto da pasta resultante.
    """
    base = Path(__file__).resolve().parent / "data"
    base.mkdir(parents=True, exist_ok=True)

    if subpasta:
        destino = base / subpasta
        destino.mkdir(parents=True, exist_ok=True)
        return destino

    return base


def localizar_arquivo(pasta: Path, nome_arquivo: str) -> Path | None:
    """Encontra um arquivo que corresponde ao padrão especificado na pasta."""

    log(f"[ARQUIVO] Buscando o arquivo: {nome_arquivo}")
    for arquivo in pasta.glob("*.xls"):
        if re.match(nome_arquivo, arquivo.name):
            log(f"[ARQUIVO] Encontrado o: {arquivo.name}")
            return arquivo
    log("[ARQUIVO] Nenhum arquivo encontrado.")

    return None


def ajustar_formula_linha(formula: str, linha_origem: int, linha_destino: int) -> str:
    """
    Ajusta referências de linha relativas em uma fórmula ao copiar para outra linha.
    Ex: '=A1+D1*C1' da linha 1 para 10 → '=A10+D10*C10'
    """

    def repl(match):
        col, row = match.groups()
        row = int(row)
        nova_linha = row - linha_origem + linha_destino
        return f"{col}{nova_linha}"

    pattern = re.compile(r"([A-Z]+)(\d+)")
    return pattern.sub(repl, formula)


def copiar_linha_com_formula(
    ws: Worksheet,
    linha_origem: int,
    linha_destino: int,
    colunas: list[int] = None,
    colunas_extras: list[str] = None,
    ajustar_formulas: bool = True,  # novo parâmetro
):
    """
    Copia uma linha para outra linha dentro da mesma planilha,
    mantendo formatação e, opcionalmente, ajustando fórmulas relativas.

    colunas: lista de índices de colunas para copiar. Se None, copia todas.
    colunas_extras: lista de letras de colunas adicionais a copiar.
    ajustar_formulas: se False, copia valores de fórmulas como estão.
    """
    if colunas is None:
        colunas = list(range(1, ws.max_column + 1))

    # Adiciona colunas extras, convertendo de letra para índice
    if colunas_extras:
        for letra in colunas_extras:
            idx = column_index_from_string(letra)
            if idx not in colunas:
                colunas.append(idx)
        colunas.sort()

    for col in colunas:
        cel_origem = ws.cell(row=linha_origem, column=col)
        cel_dest = ws.cell(row=linha_destino, column=col)

        # Copiar valor ou ajustar fórmula
        if ajustar_formulas and cel_origem.data_type == "f":
            cel_dest.value = ajustar_formula_linha(
                cel_origem.value, linha_origem, linha_destino
            )
        else:
            cel_dest.value = cel_origem.value

        # Copiar estilos
        if cel_origem.has_style:
            cel_dest.font = copy(cel_origem.font)
            cel_dest.border = copy(cel_origem.border)
            cel_dest.fill = copy(cel_origem.fill)
            cel_dest.number_format = copy(cel_origem.number_format)
            cel_dest.protection = copy(cel_origem.protection)
            cel_dest.alignment = copy(cel_origem.alignment)


def filtrar_linhas(
    ws: Worksheet,
    col_status: int,
    incluir: list[str] = None,
    excluir: list[str] = None,
    linha_inicial: int = 2,
    linha_final: int = None,
) -> list[int]:
    """
    Filtra linhas de uma planilha com base em uma coluna (status)
    e retorna uma lista com os índices das linhas que devem ser
    consideradas..
    """

    if linha_final is None:
        linha_final = ws.max_row

    # Normalizar listas de comparação
    incluir_norm = [s.strip().lower() for s in incluir] if incluir else None
    excluir_norm = [s.strip().lower() for s in excluir] if excluir else None

    linhas = []
    for row in range(linha_inicial, linha_final + 1):
        status = ws.cell(row=row, column=col_status).value
        status_val = str(status).strip().lower() if status else ""

        if incluir_norm and status_val not in incluir_norm:
            continue
        if excluir_norm and status_val in excluir_norm:
            continue

        linhas.append(row)

    log(f"[RELATÓRIO] Linhas filtradas para aba '{ws.title}': {len(linhas)}")

    return linhas


def obter_ultima_linha_com_dados(ws, coluna_chave: int):
    """Retorna a última linha da planilha que contém dados na coluna especificada."""
    return max(
        row
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=coluna_chave).value is not None
    )


def deletar_linhas(sheet, linhas, log_prefix="[TRATAMENTO]"):
    """Deleta múltiplas linhas em blocos consecutivos, reduzindo chamadas COM."""
    if not linhas:
        return

    linhas = sorted(set(linhas))  # evita duplicados
    blocos = []
    inicio = fim = linhas[0]

    for l in linhas[1:]:
        if l == fim + 1:
            fim = l
        else:
            blocos.append((inicio, fim))
            inicio = fim = l
    blocos.append((inicio, fim))

    try:
        for inicio, fim in reversed(blocos):
            if inicio == fim:
                sheet.Rows(inicio).Delete()
            else:
                sheet.Rows(f"{inicio}:{fim}").Delete()
        log(f"{log_prefix} Linhas removidas: {linhas}")
    except Exception as e:
        log(f"{log_prefix} Erro ao remover linhas {linhas}: {e}")


def salvar_excel(workbook, caminho: Path):
    """
    Salva um workbook do Excel em um arquivo.

    :param workbook: um Workbook do Excel
    :param caminho: um Path com o caminho para o arquivo a ser salvo
    """
    with log_tempo("[ARQUIVO] Salvar arquivo"):
        workbook.SaveAs(str(caminho), FileFormat=51)
        log(f"[ARQUIVO] Arquivo salvo como: {caminho.name}")


def limpar_uploads(pasta_base: Path) -> int:
    """Remove todos os arquivos .xlsx da pasta uploads.

    :param pasta_base: pasta base do projeto
    :return: quantidade de arquivos removidos
    """
    pasta = pasta_base / "uploads"
    if not pasta.exists():
        log("[DIRETORIO] Pasta uploads não encontrada")
        return 0

    removidos = 0
    for arq in pasta.glob("*.xlsx"):
        try:
            arq.unlink()
            log(f"[DIRETORIO] Removido: {arq.name}")
            removidos += 1
        except Exception as e:
            log(f"[DIRETORIO] Erro ao remover {arq.name}: {str(e)}")

    log(f"[DIRETORIO] Total removido: {removidos}")
    return removidos

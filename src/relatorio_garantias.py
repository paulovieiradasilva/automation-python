from pathlib import Path
from openpyxl import load_workbook

from config import MAPEAMENTO_COLUNAS, COLUNAS_RELATORIO
from utils import (
    log,
    log_tempo,
    copiar_linha_com_formula,
    filtrar_linhas,
    obter_ultima_linha_com_dados,
)


def abrir_planilhas():
    """Abre planilha de origem e destino, retornando workbooks e worksheets."""

    # Diretório onde os arquivos estao
    dir_base = Path(__file__).resolve().parent / "data"

    # Onde vamos filtrar os dados, para copiar na planilha de destino
    # Filtros
    # - (Resolvidos e Fechados) vão para abra - [Resolvidos-Fechados]
    # - Diferentes de (Resolvidos e Fechados) vão para RI - [RI]
    wb_origem_filtros = load_workbook(
        dir_base / "uploads" / "Filtro Incidentes (Jira).xlsx"
    )
    ws_origem_filtros = wb_origem_filtros.active

    # Planilha de destino - [Relatório]
    wb_destino_relatorio = load_workbook(
        dir_base / "Relatório Incidentes_Garantia_Projetos_v5.xlsx"
    )
    ws_destino_resolvidos_fechados = wb_destino_relatorio["Resolvidos-Fechados"]
    ws_destino_ri = wb_destino_relatorio["RI"]
    ws_destino_projetos = wb_destino_relatorio["Projetos"]

    # Planilha de origem - [Extração de Projetos]
    wb_origem_projetos = load_workbook(dir_base / "uploads" / "Projetos (Jira).xlsx")
    ws_origem_projetos = wb_origem_projetos.active

    return (
        wb_origem_filtros,
        ws_origem_filtros,
        wb_destino_relatorio,
        ws_destino_resolvidos_fechados,
        ws_destino_ri,
        wb_origem_projetos,
        ws_origem_projetos,
        ws_destino_projetos,
    )


def preparar_mapeamento(ws_origem, ws_destino, mapa_colunas):
    """
    Prepara índices e colunas para cópia entre origem e destino.
    Exige que seja passado um mapeamento customizado de colunas.
    """
    cabecalhos_origem = [cell.value for cell in ws_origem[1]]
    cabecalhos_destino = [cell.value for cell in ws_destino[1]]

    indices_origem = {name: idx + 1 for idx, name in enumerate(cabecalhos_origem)}
    indices_destino = {name: idx + 1 for idx, name in enumerate(cabecalhos_destino)}

    # Não precisa mais de fallback; mapa_colunas é obrigatório
    colunas_para_copiar = [
        indices_destino[dest]
        for dest in mapa_colunas.values()
        if dest in indices_destino
    ]

    return mapa_colunas, indices_origem, indices_destino, colunas_para_copiar


def preparar_mapeamento_simples(ws_origem, ws_destino):
    """Prepara índices e colunas para cópia entre origem e destino."""

    # Cria os cabeçalhos
    cabecalhos_origem = [cell.value for cell in ws_origem[1]]
    cabecalhos_destino = [cell.value for cell in ws_destino[1]]

    # Cria os indices
    indices_origem = {name: idx + 1 for idx, name in enumerate(cabecalhos_origem)}
    indices_destino = {name: idx + 1 for idx, name in enumerate(cabecalhos_destino)}

    # Mapeamento entre colunas de origem e destino
    mapa_colunas = {col: col for col in cabecalhos_destino if col in cabecalhos_origem}

    # Colunas do destino que devem ser copiadas
    colunas_para_copiar = [
        indices_destino[dest] for dest in mapa_colunas.keys() if dest in indices_destino
    ]

    return mapa_colunas, indices_origem, indices_destino, colunas_para_copiar


def preparar_destino(ws_destino, linha_modelo: int = 2):
    """Limpa a planilha de destino, mantendo a linha modelo."""

    if ws_destino.max_row > linha_modelo:
        ws_destino.delete_rows(linha_modelo + 1, ws_destino.max_row - linha_modelo)
    log(f"Após limpeza, '{ws_destino.title}' tem {ws_destino.max_row} linhas")


def copiar_para_aba(
    ws_destino,
    linha_modelo,
    linhas_origem,
    linha_destino,
    ws_origem,
    mapa_colunas,
    indice_origem,
    indice_destino,
    colunas_para_copiar,
    ajustar_formulas: bool = True,
    colunas_extras: list[str] = None,
):
    """Copia linhas da origem para a aba de destino mantendo formatação e fórmulas."""
    for linha_origem in linhas_origem:
        copiar_linha_com_formula(
            ws_destino,
            linha_origem=linha_modelo,
            linha_destino=linha_destino,
            colunas=colunas_para_copiar,
            colunas_extras=colunas_extras if colunas_extras else [],
            ajustar_formulas=ajustar_formulas,
        )

        for coluna_destino, coluna_origem in mapa_colunas.items():
            idx_dest = indice_destino.get(coluna_destino)
            idx_origem = indice_origem.get(coluna_origem)
            if idx_dest is None or idx_origem is None:
                continue

            valor_origem = ws_origem.cell(row=linha_origem, column=idx_origem).value
            ws_destino.cell(row=linha_destino, column=idx_dest).value = valor_origem

        linha_destino += 1

    return linha_destino - 2


def obter_ultima_linha(ws, coluna_chave):
    """
    Retorna o índice da última linha com dados em uma coluna chave.
    Args: ws: Worksheet (aba do Excel) coluna_chave (str): Nome do cabeçalho da coluna usada como referência.
    Returns: int: Número da última linha com dados.
    """
    cabecalho = [cell.value for cell in ws[1]]
    if coluna_chave not in cabecalho:
        raise ValueError(
            f"Coluna '{coluna_chave}' não encontrada no cabeçalho da aba {ws.title}"
        )

    idx_col = cabecalho.index(coluna_chave) + 1
    return obter_ultima_linha_com_dados(ws, idx_col)


def processar_rf(ws_origem, ws_destino):

    # Número da Linha Modelo.
    nun_linha = 145

    # Limpar aba de destino
    preparar_destino(ws_destino, linha_modelo=nun_linha)

    # Obter a última linha
    ultima_linha = obter_ultima_linha(ws_origem, "Chave")

    mapa_colunas, indice_origem, indice_destino, colunas_para_copiar = (
        preparar_mapeamento(ws_origem, ws_destino, MAPEAMENTO_COLUNAS)
    )

    # Copiar para RF (filtrando status = 'Resolvido' e 'Finalizado')
    linhas_para_RF = filtrar_linhas(
        ws_origem,
        indice_origem["Situação"],
        incluir=["Resolvido", "Finalizado"],
        linha_inicial=2,
        linha_final=ultima_linha,
    )
    copiar_para_aba(
        ws_destino,
        linha_modelo=nun_linha,
        linhas_origem=linhas_para_RF,
        linha_destino=nun_linha,
        ws_origem=ws_origem,
        mapa_colunas=mapa_colunas,
        indice_origem=indice_origem,
        indice_destino=indice_destino,
        colunas_para_copiar=colunas_para_copiar,
        colunas_extras=COLUNAS_RELATORIO,
    )


def processar_ri(ws_origem, ws_destino):

    # Número da Linha Modelo.
    nun_linha = 2

    # Limpar aba de destino
    preparar_destino(ws_destino)

    # Obter a última linha
    ultima_linha = obter_ultima_linha(ws_origem, "Chave")

    # Preparar mapeamento entre origem e destino
    mapa_colunas, indice_origem, indice_destino, colunas_para_copiar = (
        preparar_mapeamento(ws_origem, ws_destino, MAPEAMENTO_COLUNAS)
    )

    linhas_para_RI = filtrar_linhas(
        ws_origem,
        indice_origem["Situação"],
        excluir=["Resolvido", "Finalizado"],
        linha_inicial=2,
        linha_final=ultima_linha,
    )
    copiar_para_aba(
        ws_destino,
        linha_modelo=nun_linha,
        linhas_origem=linhas_para_RI,
        linha_destino=nun_linha,
        ws_origem=ws_origem,
        mapa_colunas=mapa_colunas,
        indice_origem=indice_origem,
        indice_destino=indice_destino,
        colunas_para_copiar=colunas_para_copiar,
        colunas_extras=COLUNAS_RELATORIO,
    )


def processar_projetos(ws_origem, ws_destino):

    # Número da Linha Modelo.
    nun_linha = 2

    # Limpar aba de destino
    preparar_destino(ws_destino)

    # Obter a última linha
    ultima_linha = obter_ultima_linha(ws_origem, "Chave")

    # prepara mapeamento entre origem e destino
    mapa_colunas, indice_origem, indice_destino, colunas_para_copiar = (
        preparar_mapeamento_simples(ws_origem, ws_destino)
    )

    linhas = list(range(2, ultima_linha + 1))
    total_copiados = copiar_para_aba(
        ws_destino,
        linha_modelo=nun_linha,
        linhas_origem=linhas,
        linha_destino=nun_linha,
        ws_origem=ws_origem,
        mapa_colunas=mapa_colunas,
        indice_origem=indice_origem,
        indice_destino=indice_destino,
        colunas_para_copiar=colunas_para_copiar,
        ajustar_formulas=False,
    )
    log(f"Total de registros copiados para aba {ws_destino.title}: {total_copiados}")


def main():
    with log_tempo("[RELATÓRIOS] ~ Processamento"):
        # Diretório onde os arquivos estao
        dir_base = Path(__file__).resolve().parent / "data"

        # Abrir planilhas
        (
            wb_origem_filtros,
            ws_origem_filtros,
            wb_destino_relatorio,
            ws_destino_relatorio,
            ws_destino_ri,
            wb_origem_projetos,
            ws_origem_projetos,
            ws_destino_projetos,
        ) = abrir_planilhas()

        with log_tempo("Copia de projetos"):
            # [Projetos]
            processar_projetos(ws_origem_projetos, ws_destino_projetos)

        with log_tempo("Copia de RI"):
            # [RI - Chamados Abertos]
            processar_ri(ws_origem_filtros, ws_destino_ri)

        with log_tempo("Copia de Resolvidos e Fechados"):
            # [Resolvidos e Fechados]
            processar_rf(ws_origem_filtros, ws_destino_relatorio)

    # Salvar planilha
    wb_destino_relatorio.save(
        dir_base / "Relatório Incidentes_Garantia_Projetos_v5_(2).xlsx"
    )
    log("Relatório salvo com sucesso.")

    # Fechar os workbooks
    wb_origem_filtros.close()
    wb_destino_relatorio.close()
    wb_origem_projetos.close()


if __name__ == "__main__":
    main()

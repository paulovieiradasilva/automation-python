from pathlib import Path
import win32com.client as win32
from openpyxl import load_workbook
from utils import log, log_tempo, localizar_arquivo, salvar_excel


def tratar_arquivo_xlsx(caminho: Path):
    wb = load_workbook(caminho)
    ws = wb.active

    ultima_linha = ws.max_row
    valores_ultima_linha = [cell.value or "" for cell in ws[ultima_linha]]

    with log_tempo("Remover última linha"):
        if any("Gerado em" in str(valor) for valor in valores_ultima_linha):
            log(f"A última linha ({ultima_linha}) contém 'Gerado em'. Removendo...")
            ws.delete_rows(ultima_linha)
        else:
            log(
                f"A última linha ({ultima_linha}) NÃO contém 'Gerado em'. Nada a remover."
            )
        wb.save(caminho)
        log("Alterações salvas com openpyxl.")

    with log_tempo("Abrir Excel para ajustes finais"):
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        workbook = excel.Workbooks.Open(str(caminho))
        sheet = workbook.Sheets(1)

        with log_tempo("Desativar quebra de texto"):
            sheet.UsedRange.WrapText = False
            log("Quebra de texto desativada.")

        with log_tempo("Remover imagens"):
            total_imagens = 0
            for shape in sheet.Shapes:
                if shape.Type == 13:  # msoPicture
                    shape.Delete()
                    total_imagens += 1
            log(f"Total de imagens removidas: {total_imagens}")

        with log_tempo("Remover linhas 1-3"):
            sheet.Rows("1:3").Delete()
            log("Linhas 1, 2 e 3 removidas.")

        workbook.Close(SaveChanges=True)
        excel.Quit()
        log("Excel fechado e ajustes finais aplicados com sucesso.")


def xls_para_xlsx(
    pasta: Path, xls: str, xlsx: str, deletar_xls: bool = True
) -> Path | None:
    """
    Converte um arquivo .xls para .xlsx mantendo formatação.
    Apenas faz uma ação: abrir e salvar como .xlsx.
    """
    with log_tempo("Conversão XLS para XLSX"):
        caminho_xls = pasta / xls
        caminho_xlsx = pasta / "uploads" / xlsx
        caminho_xlsx.parent.mkdir(parents=True, exist_ok=True)

        if not caminho_xls.exists():
            raise FileNotFoundError(f"O arquivo não foi encontrado: {caminho_xls}")

        if caminho_xlsx.exists():
            caminho_xlsx.unlink()

        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        workbook = None

        try:
            # Abrir arquivo .xls
            workbook = excel.Workbooks.Open(str(caminho_xls))

            # Salvar como .xlsx
            salvar_excel(workbook, caminho_xlsx)

            # Deletar o .xls original se necessário
            if deletar_xls:
                caminho_xls.unlink()
                log(f"Arquivo .xls removido: {caminho_xls}")

            return caminho_xlsx

        except Exception as e:
            log(f"Falha na conversão: {e}")
            return None

        finally:
            if workbook:
                workbook.Close(SaveChanges=False)
            excel.Quit()


def processar_arquivos_xls(folder_data: Path, arquivos_info: list[dict], del_xls: bool):
    for regex, novo_nome in arquivos_info:
        xls = localizar_arquivo(folder_data, regex)
        if xls:
            xlsx = xls_para_xlsx(
                pasta=folder_data,
                xls=xls.name,
                xlsx=novo_nome,
                deletar_xls=del_xls,
            )
            # Tratar o .xlsx
            tratar_arquivo_xlsx(xlsx)
        else:
            log(f"Nenhum arquivo encontrado para padrão: {regex}")


def main():
    # Caminho da pasta 'data'.
    folder_data = Path(__file__).parent / "data"

    print(f" # ~ Processando arquivos na pasta: {folder_data}")

    # Mapeamento dos arquivos .xls e .xlsx
    mapeamento = [
        (
            r"Filtro Incidentes - Garantia de Projetos \(Jira\).*\.xls",
            "Filtro Incidentes (Jira).xlsx",
        ),
        (r"Projetos \(Jira\).*\.xls", "Projetos (Jira).xlsx"),
        (r"Defeitos SKY AD \(Jira\).*\.xls", "Defeitos SKY AD (Jira).xlsx"),
        (r"Project Room \(Jira\).*\.xls", "Project Room (Jira).xlsx"),
        (r"Relatório RM \(Jira\).*\.xls", "Relatório RM (Jira).xlsx"),
    ]

    with log_tempo("Processamento de arquivos"):
        # Processar os arquivos .xls
        processar_arquivos_xls(folder_data, mapeamento, del_xls=False)


if __name__ == "__main__":
    main()
